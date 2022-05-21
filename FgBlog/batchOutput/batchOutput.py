import pymysql
import xlsxwriter
import xlwt
import datetime
import time
import datetime
from hr_manage_db import models
from django.forms.models import model_to_dict
from openpyxl import Workbook, load_workbook
import paramiko
import json
fileType = 'test'
# workBook1 = xlsxwriter.Workbook(f'../downloadFile/{datetime.date.today()}-{fileType}.xlsx')




def getFileList(excelType):
    if excelType == 'ApplicantInfo':
        excelTypeName = '候选人管理'
    if excelType == 'ProjectInfo':
        excelTypeName = '项目信息'
    if excelType == 'ProjectStatusInfo':
        excelTypeName = '项目满足度'
    if excelType == 'RecruitmentInfo':
        excelTypeName = '招聘需求'
    # 获取transport传输实例, sftp服务器 ip + 端口号
    tran = paramiko.Transport(('139.9.160.24', 22))
    # 连接ssh服务器, user + password
    tran.connect(username='root', password='lunchFeng@07177')
    # 获取sftp实例
    sftp = paramiko.SFTPClient.from_transport(tran)
    path = '/tmp/hr_manage/FgBlog/dist/static/downloadFile/'
    filesAttr = sftp.listdir_attr(path)
    filesAttr.sort(key=lambda f: f.filename)
    filesAttr = filesAttr[::-1]
    fileList = []
    for fileAttr in filesAttr:
        filename = fileAttr.filename
        print(fileAttr.longname)
        if excelType in filename or excelTypeName in filename:
            fileList.append({'path': 'static/downloadFile/' + filename, 'name': filename})
    return fileList


def newDownloadExcel(excelType, **kwargs):
    if excelType == 'ApplicantInfo':
        fileName = 'applicant_info'
    if excelType == 'ProjectInfo':
        fileName = 'project_info'
    if excelType == 'ProjectStatusInfo':
        fileName = 'project_status_info'
    if excelType == 'RecruitmentInfo':
        fileName = 'recruitment_info'
    m = readJson(fileName)
    print(m)
    if excelType == 'ApplicantInfo':
        dbTarget = models.ApplicantInfo
        target = models.ApplicantInfo
        excelTypeName = '候选人管理'
    if excelType == 'ProjectInfo':
        dbTarget = models.ProjectInfo
        target = models.ProjectInfo
        excelTypeName = '项目信息'
    if excelType == 'ProjectStatusInfo':
        dbTarget = models.ProjectStatusInfo
        target = models.ProjectStatusInfo
        excelTypeName = '项目满足度'
    if excelType == 'RecruitmentInfo':
        dbTarget = models.RecruitmentInfo
        target = models.RecruitmentInfo
        excelTypeName = '招聘需求'
    try:
        dbTarget = dbTarget._meta.get_fields()
        # wbk = xlwt.Workbook()
        now = str(datetime.datetime.now()).split(".")[0]
        filename = f'{now}-{excelTypeName}.xlsx'
        wb = Workbook()
        ws = wb['Sheet']
        # sheet = wbk.add_sheet('Sheet1', cell_overwrite_ok=True)
        keyList = []
        for i in range(len(dbTarget)):
            query = dbTarget[i]
            name = query.name
            print(query.name)
            keyList.append(name)
            # sheet.write(0, i, name)
            # ws.cell(1, i+1).value = name
        i= 0
        for item in m:
            ws.cell(1, i+1).value = m[item]
            i+=1
        target = target.objects.all()
        try:
            if kwargs['region'] and kwargs['region']!='all':
                print('############') # 过滤可选数据
                regionList = kwargs['region'].split('|')
                targetList = []
                for i in regionList:
                    targetList.append(target.filter(region=i))
                print(regionList)
                print(targetList)
                target = ''
                for i in targetList:
                    if target == '':
                        target = i
                    else:
                        target = target | i
        except:
            print('没传地域')
        try:
            if kwargs['selectDate']:
                date = kwargs['selectDate']
                print(date)
                target = target.filter(date=date)
        except:
            print('没传日期')
        for i in range(len(target)):
            targetI = model_to_dict(target[i])
            for j in range(len(keyList)):
                value = keyList[j]
                value = targetI[value]
                # print(str(value))
                # sheet.write(i + 1, j, str(value))
                ws.cell(i + 2, j+1).value = str(value)
        # wbk.save(f'FgBlog/dist/static/downloadFile/{filename}')
        wb.save(f'FgBlog/dist/static/downloadFile/{filename}')
        if excelType == 'ProjectStatusInfo':
            # 将从数据库获取的数据填入月度分析demo后，保存为新excel
            # 读取需要是xlsx格式的文件
            # 是否可以根据筛选的月份，只下载当前月份的数据呀？
            # 打开的页面是否需要关闭？会不会内存泄漏？
            print(111)
            wb_demo = load_workbook('FgBlog/dist/static/downloadFile/海思人力供给进展月度汇总分析-demo.xlsx')
            print(222)
            wb_src = load_workbook(f'FgBlog/dist/static/downloadFile/{filename}')
            print(333)
            ws_demo = wb_demo['月度汇总页']
            ws_src = wb_src['Sheet']
            print('插入当月数据')
            for i in range(ws_src.max_row):
                for j in range(13):
                    # 不能识别为数字,等待数据库处理后，再看下效果
                    ws_demo.cell(i + 3, j + 1).value = ws_src.cell(i + 2, j + 3).value
            # 写入当前月份
            print('写入当前月份')
            ws_demo['b1'].value = ws_src['b2'].value
            # 保存为新的文件
            print('保存为新的文件')
            filename = f'after-{filename}'
            wb_demo.save(f'FgBlog/dist/static/downloadFile/{filename}')
            print('###############\n###############')
        return {'path': 'static/downloadFile/' + filename, 'name': filename}
    except Exception as e:
        print(str(e))
        print('error', excelType)


def readJson(type):
    print(type)
    obj = json.load(open(f'/tmp/hr_manage/FgBlog/tableJson/{type}.json', 'r', encoding='utf-8'))
    # for i in obj:
    #     print(i)
    #     print(obj[i])
    return obj